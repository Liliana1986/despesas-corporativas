"""
Geração de relatório Excel com todas as despesas e resultado da reconciliação.
"""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


VERDE = "C6EFCE"
VERMELHO = "FFC7CE"
AMARELO = "FFEB9C"
CINZA = "D9D9D9"
AZUL_HEADER = "1F4E79"


def generate_report(faturas_df: pd.DataFrame, reconciliacao: dict | None = None) -> bytes:
    """Gera Excel com aba de despesas e, opcionalmente, abas de reconciliação."""
    wb = Workbook()

    _create_expenses_sheet(wb, faturas_df)

    if reconciliacao:
        _create_reconciliation_sheet(wb, reconciliacao)
        _create_missing_sheet(wb, reconciliacao)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _create_expenses_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Despesas Extraídas"

    columns = [
        ("Colaborador", 20),
        ("Ficheiro", 30),
        ("Fornecedor", 25),
        ("Data", 12),
        ("Descrição", 35),
        ("Valor (€)", 12),
        ("Nº Fatura", 18),
        ("Estado", 20),
    ]

    _write_header(ws, [c[0] for c in columns], AZUL_HEADER)
    _set_column_widths(ws, columns)

    col_map = {
        "Colaborador": "colaborador",
        "Ficheiro": "ficheiro",
        "Fornecedor": "fornecedor",
        "Data": "data",
        "Descrição": "descricao",
        "Valor (€)": "valor_total",
        "Nº Fatura": "numero_fatura",
        "Estado": "erro",
    }

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, (header, _) in enumerate(columns, start=1):
            key = col_map.get(header, "")
            value = row.get(key, "")
            if value is None or str(value) == "nan":
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")

            if header == "Valor (€)" and value:
                try:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00 "€"'
                except (ValueError, TypeError):
                    pass

            erro = row.get("erro", "")
            if erro and str(erro) != "nan":
                cell.fill = PatternFill(fill_type="solid", fgColor=AMARELO)
            else:
                fill = VERDE if row_idx % 2 == 0 else "FFFFFF"
                cell.fill = PatternFill(fill_type="solid", fgColor=fill)

        ws.row_dimensions[row_idx].height = 18

    _add_border(ws, len(df) + 1, len(columns))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _create_reconciliation_sheet(wb: Workbook, reconciliacao: dict):
    ws = wb.create_sheet("Reconciliação")

    conciliadas = reconciliacao.get("conciliadas", pd.DataFrame())
    sem_doc = reconciliacao.get("sem_documento", pd.DataFrame())
    sem_ext = reconciliacao.get("sem_extrato", pd.DataFrame())

    # Resumo
    ws["A1"] = "RESUMO DA RECONCILIAÇÃO"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=AZUL_HEADER)
    ws.merge_cells("A1:D1")

    resumo = [
        ("Despesas conciliadas", len(conciliadas), VERDE),
        ("Movimentos sem documento", len(sem_doc), VERMELHO),
        ("Documentos sem extrato", len(sem_ext), AMARELO),
        ("Total documentos processados", len(conciliadas) + len(sem_ext), CINZA),
    ]

    for i, (label, count, color) in enumerate(resumo, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=count)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.font = Font(bold=True)

    # Tabela conciliadas
    start_row = 8
    ws.cell(row=start_row, column=1, value="DESPESAS CONCILIADAS").font = Font(bold=True, size=12)
    start_row += 1

    if not conciliadas.empty:
        cols = ["colaborador", "fornecedor", "data_fatura", "valor_fatura", "descricao_extrato", "valor_extrato", "estado"]
        headers = ["Colaborador", "Fornecedor", "Data Fatura", "Valor Fatura", "Descrição Extrato", "Valor Extrato", "Estado"]
        _write_header(ws, headers, "2E7D32", start_row=start_row)
        for r, (_, row) in enumerate(conciliadas.iterrows(), start=start_row + 1):
            for c, col in enumerate(cols, start=1):
                cell = ws.cell(row=r, column=c, value=str(row.get(col, "")))
                cell.fill = PatternFill(fill_type="solid", fgColor=VERDE)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 25


def _create_missing_sheet(wb: Workbook, reconciliacao: dict):
    ws = wb.create_sheet("Em Falta")

    sem_doc = reconciliacao.get("sem_documento", pd.DataFrame())
    sem_ext = reconciliacao.get("sem_extrato", pd.DataFrame())

    ws["A1"] = "MOVIMENTOS SEM DOCUMENTO"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="C62828")
    ws.merge_cells("A1:D1")

    row = 2
    if not sem_doc.empty:
        headers = ["Data Extrato", "Descrição", "Valor", "Estado"]
        _write_header(ws, headers, "C62828", start_row=row)
        row += 1
        for _, r in sem_doc.iterrows():
            ws.cell(row=row, column=1, value=str(r.get("data_extrato", "")))
            ws.cell(row=row, column=2, value=str(r.get("descricao_extrato", "")))
            ws.cell(row=row, column=3, value=str(r.get("valor_extrato", "")))
            ws.cell(row=row, column=4, value=str(r.get("estado", "")))
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = PatternFill(fill_type="solid", fgColor=VERMELHO)
            row += 1

    row += 2
    ws.cell(row=row, column=1, value="DOCUMENTOS SEM CORRESPONDÊNCIA NO EXTRATO")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(fill_type="solid", fgColor="E65100")
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    if not sem_ext.empty:
        headers = ["Colaborador", "Ficheiro", "Fornecedor", "Data Fatura", "Valor", "Estado"]
        _write_header(ws, headers, "E65100", start_row=row)
        row += 1
        for _, r in sem_ext.iterrows():
            for c, key in enumerate(["colaborador", "ficheiro", "fornecedor", "data_fatura", "valor_fatura", "estado"], start=1):
                cell = ws.cell(row=row, column=c, value=str(r.get(key, "")))
                cell.fill = PatternFill(fill_type="solid", fgColor=AMARELO)
            row += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22


def _write_header(ws, headers: list, color: str, start_row: int = 1):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(ws, columns: list):
    for col_idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_border(ws, max_row: int, max_col: int):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
